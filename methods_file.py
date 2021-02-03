import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import re
from ebooklib import epub
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as Pq
from breadability.readable import Article
import random
import base64
from datetime import datetime, timedelta, date
from difflib import SequenceMatcher
from time import sleep
import functools
from PyPDF4 import PdfFileReader, PdfFileWriter, PdfFileMerger
from PIL import Image, ImageDraw, ImageFont
from urllib.parse import unquote
from collections import namedtuple


Item_Entry = namedtuple(typename='Item_Entry',
                        field_names=['HTML_Address', 'Pdf_Address', 'Article_Title', 'List_Index'])


def http_error(func):
    num_retries = 3

    @functools.wraps(func)
    def wrapper(*a, **kw):
        # lines = inspect.stack(context=2)[1].code_context
        # decorated = any(line.startswith('@') for line in lines)
        # print(func.__name__, 'was decorated with "@decorate":', decorated)
        sleep_interval = 4
        for i in range(num_retries):
            try:
                return func(*a, **kw)
            except Exception as e:
                if i < num_retries - 1:
                    sleep(sleep_interval)
                    # print("Oops!", e.__class__, "occurred.")
                    sleep_interval = min(2 * sleep_interval, 60)
                else:
                    print('Fatal error... exceeded retries')
                    print(f'Shit happened in {func.__name__}\n\twith argument {a[1]}\n\terror is: {e.__class__}'
                          f'\n\tretrying in 4 seconds')
                    print(e)
                    raise e
    return wrapper


class GetResourceMethods:
    phantomjs_exe_path = "phantomjs.exe"
    pdf_js_path = "papersize.js"
    opinion_articles_headers = ['opinion', 'wsj.com', 'theguardian.com/commentisfree/', 'globaltimes.cn/',
                                'livemint.com/', 'dawn.com/']
    regex_search = 'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
    final_pdf_path = "D:\\UPSC\\UPSC 2020\\Newspaper and others\\Hindu Mint and IE"
    final_file_name_pdf = f'{str(date.today())}_Compilation_by_FatPanda.pdf'
    epub_file_name = str(date.today()) + "_Indian_Express.epub"
    explained_list, opinion_list, other_list, economist_list, \
        opinion_chapters, other_chapters, explained_chapters, economist_chapters = ([] for gu in range(8))
    indian_express_epub = epub.EpubBook()

    def __init__(self):
        self.decision_dict = {
            'https://www.theguardian.com/': self.parse_guradian_nytimes_globaltimes_url,
            'https://www.wsj.com/': self.parse_wsj_url,
            'https://www.thehindu.com/': self.parse_hindu_faster,
            'https://www.washingtonpost.com/': self.parse_wp_url_ampway,
            'https://indianexpress.com/': self.parse_indian_express_faster,
            'https://www.livemint.com/': self.parse_livemint_url,
            'https://www.globaltimes.cn/': self.parse_guradian_nytimes_globaltimes_url,
            'https://www.nytimes.com/': self.parse_guradian_nytimes_globaltimes_url,
            'https://www.taipeitimes.com/': self.parse_taipei_times,
            'https://www.downtoearth.org.in/': self.parse_dte,
            'https://hongkongfp.com/': self.parse_hkfp,
            'https://www.epw.in/': self.parse_epw_non_outline,
            'https://www.economist.com/': self.parse_economist,
            'https://www.sapiens.org/': self.parse_sapiens,
            'https://perspectivesinanthropology.com/': self.parse_perspective_anthro,
            'https://www.insightsonindia.com/': self.parse_insights_daily_non_outline,
            'http://www.thepeoplehistory.com/': self.parse_history_articles,
            'https://www.indianage.com/': self.parse_history_articles,
            'https://www.dawn.com/': self.parse_dawn,
            'https://medium.com/': self.parse_outline_url
        }

    def update_lists(self, html_file_add, pdf_file_add, art_url, art_title, content_index):
        # print(html_file_add,pdf_file_add,art_url,art_title)
        if 'explained' in art_url:
            self.explained_list.append(Item_Entry._make([html_file_add, pdf_file_add, art_title, content_index]))
            # self.explained_list.append([html_file_add, pdf_file_add, art_title])
        elif any(op_art in art_url for op_art in self.opinion_articles_headers):
            self.opinion_list.append(Item_Entry._make([html_file_add, pdf_file_add, art_title, content_index]))
        elif 'economist.com/' in art_url or 'epw.in/' in art_url:
            self.economist_list.append(Item_Entry._make([html_file_add, pdf_file_add, art_title, content_index]))
        else:
            self.other_list.append(Item_Entry._make([html_file_add, pdf_file_add, art_title, content_index]))

    def add_print_css(self):
        """

        :return: html string after adding style sheet
        """
        with open('print.css') as print_css:
            out_html = print_css.read()
        return out_html

    def get_random_response(self, url):
        """

        :param url: requests.get will be run on this
        :return: http response
        """
        user_agent_list = [
            # Chrome
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 5.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 '
            'Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 '
            'Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 '
            'Safari/537.36',
            # Firefox
            'Mozilla/4.0 (compatible; MSIE 9.0; Windows NT 6.1)',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0)',
            'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (Windows NT 6.2; WOW64; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0; Trident/5.0)',
            'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0)',
            'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0)',
            'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; .NET CLR 2.0.50727; '
            '.NET CLR 3.0.4506.2152; .NET CLR 3.5.30729)',
            # Google Bot
            'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)',
            # Bing bot
            'Mozilla/5.0 (compatible; bingbot/2.0; +http://www.bing.com/bingbot.htm)',
            # Yahoo! bot
            'Mozilla/5.0 (compatible; Yahoo! Slurp; http://help.yahoo.com/help/us/ysearch/slurp)'
        ]
        user_agent = random.choice(user_agent_list)
        headers = {'User-Agent': user_agent}
        response = requests.get(url, headers=headers)
        return response

    @http_error
    def get_amp_url_requests(self, non_amp_url):
        # print('amp function called')
        response_amp = self.get_random_response(non_amp_url)
        if response_amp.status_code == 404:
            print('404 error take it selenium way')
            return None, None
        soup_amp = BeautifulSoup(response_amp.content, "lxml")
        title = re.sub("^\s+", '', soup_amp.title.text)
        try:
            amp_url = soup_amp.find('link', {'rel': 'amphtml'})['href']
            return amp_url, title
        except Exception as e:
            # print(f'this error should not occur on livemint, shall code this when it occurs, '
            #       f'livemint urls should have amp link {non_amp_url}')
            # print(e.__class__)
            return None, title

    @http_error
    def get_amp_url_selenium(self, non_amp_url):
        # print('selenium amp function called')
        options = Options()
        fp = webdriver.FirefoxProfile(
            'C:\\Users\\Sabyasachi\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\703g68w9.python_user')
        fp.DEFAULT_PREFERENCES['frozen']['extensions.autoDisableScopes'] = 0
        options.set_preference('extensions.enabledScopes', 15)
        # options.set_preference("browser.link.open_newwindow", 3)
        # options.set_preference("browser.link.open_newwindow.restriction", 0)
        fp.update_preferences()
        driver = webdriver.Firefox(firefox_profile=fp, options=options, executable_path='C:\\Users\\Sabyasachi\\Google Drive\\Python Projects\\Daily_Compiler\\geckodriver.exe')
        driver.get(non_amp_url)
        soup_amp = BeautifulSoup(driver.page_source, 'lxml')
        title = re.sub('^\s+', '', soup_amp.title.text)
        driver.close()
        try:
            amp_url = soup_amp.find('link', {'rel': 'amphtml'})['href']
            return amp_url, title
        except Exception:
            print(f'this error should not occur on livemint, shall code this when it occurs, '
                  f'livemint urls should have amp link {non_amp_url}')
            return None, title

    @http_error
    def parse_outline_url(self, input_url, url_index):
        regex_search = 'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        url = 'https://outline.com/' + input_url
        webdriver.DesiredCapabilities.FIREFOX['proxy'] = {
            "proxyType": 'DIRECT'
        }
        browser = webdriver.Firefox(executable_path='C:\\Users\\Sabyasachi\\Google Drive\\Python Projects\\Daily_Compiler\\geckodriver.exe')
        browser.get(url)
        try:
            element = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "yue"))
            )
        except Exception:
            pass
        html_epw = browser.page_source
        soup_epw = BeautifulSoup(html_epw, "lxml")
        for sc in soup_epw(['script']):
            sc.decompose()
        browser.close()
        raw_content = soup_epw.find("raw")  # , {"class": "yue"})
        # print(soup_epw.title.text)
        new_soup = BeautifulSoup(raw_content['content'], 'lxml')
        html1 = f'<html><head><meta charset="utf-8"><title>{soup_epw.title.text}</title>'
        html1 += self.add_print_css()
        html1 += f'</head><body><h1><a href="{input_url}">{url_index}_{soup_epw.title.text}</a></h1>'
        for child in new_soup.body.children:
            if not child.name:
                continue
            if '<figure>' in str(child):
                img_url = re.findall(pattern=regex_search, string=str(child))
                if len(img_url) >= 1:
                    img_url = img_url[0]
                    img_response = requests.get(img_url)
                    encoded_string = base64.b64encode(img_response.content).decode('ascii')
                    # print('adding image')
                    html1 += f'<p><img src="data:image/jpeg;base64,{encoded_string}"/></p>'
            else:
                html1 += str(child)
        title = re.sub('^\s+', '', soup_epw.title.text)
        return html1, title

    @http_error
    def parse_dawn(self, input_url, url_index):
        # print('dawn called')
        url_index = 'is_dummy'
        amp_url, title = self.get_amp_url_requests(input_url)
        response_dawn = self.get_random_response(amp_url)
        soup = BeautifulSoup(response_dawn.content, 'lxml')
        header = soup.find('h1', {'class': 'story__title'})
        author = soup.find('div', {'class': 'story__meta'})
        body = soup.find('div', {'class': 'story__content'})
        for unwanted in body.find_all(True, {'class': 'amp-ad-container'}):
            unwanted.decompose()
        html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        html += self.add_print_css()
        html += '</head><body>' + str(header) + str(author) + str(body) + '</body></html>'
        return html, title

    @http_error
    def parse_epw_non_outline(self, input_url, url_index):
        response_epw = self.get_random_response(input_url)
        if response_epw.status_code == 404:
            input_url = unquote(input_url)
            url_list = self.rebuild_epw(call_func=1)
            for url in url_list:
                if SequenceMatcher(None, url, input_url).ratio() > 0.95:
                    input_url = url
        response_epw = self.get_random_response(input_url)
        soup_epw = BeautifulSoup(response_epw.content, 'lxml')
        header = soup_epw.find('h1', {'id': 'page-title'}).text
        title = soup_epw.title.text
        html_epw = f'<html><head><meta charset="utf-8"><title>{soup_epw.title.text}</title>'
        html_epw += self.add_print_css()
        html_epw += f'</head><body><h1><a href="{input_url}">{url_index}_{header}</a></h1>'
        article = soup_epw.find('div', {'id': 'block-system-main'})
        article = article.findAll('div', {'class', 'content'})
        article = article[1]
        for figure in article.findAll('img'):
            img_url = figure['src']
            try:
                img_response = requests.get(img_url)
            except requests.exceptions.MissingSchema:
                # print('appending host to url')
                img_url = 'https://www.epw.in/' + img_url
                img_response = requests.get(img_url)
            encoded_string = base64.b64encode(img_response.content).decode('ascii')
            figure['src'] = f"data:image/jpeg;base64,{encoded_string}"
        html_epw += str(article) + '</body></html>'
        return html_epw, title

    @http_error
    def parse_dte(self, input_url, url_index):
        amp_url, title = self.get_amp_url_requests(input_url)
        if not amp_url and not title:
            amp_url, title = self.get_amp_url_selenium(input_url)
        if title and not amp_url:
            dte_html, title = self.parse_outline_url(input_url=input_url, url_index=url_index)
            return dte_html, title
        # print(amp_url)
        response_dte = self.get_random_response(amp_url)
        soup_dte = BeautifulSoup(response_dte.content, 'lxml')
        dte_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        dte_html += self.add_print_css()
        dte_html += "</head><body>"
        article = soup_dte.find('div', {'class', 'news-detail'})
        for ka in article.find_all('amp-img'):
            ka.decompose()
        for ku in article.find_all(True, {'class': ['captionStory', 'add-comment', 'flexible-item', 'latest-article',
                                                    'read-post-comment-div', 'donate-text']}):
            ku.decompose()
        article.find('header').decompose()
        header_tag = soup_dte.find('h1')
        new_header_tag = soup_dte.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        dte_html += str(article)
        return dte_html, title

    @http_error
    def parse_livemint_url(self, input_url, url_index):
        amp_url, title = self.get_amp_url_requests(input_url)
        if not amp_url and not title:
            amp_url, title = self.get_amp_url_selenium(input_url)
        response_livemint = self.get_random_response(amp_url)
        soup_livemint = BeautifulSoup(response_livemint.content, 'lxml')
        # driver.close()
        article = soup_livemint.find('div', {'class': 'mainSec'})
        try:
            article.find('div', {'class': ['bcrumb', 'promotion', 'share-icons-box', 'epaperPromo']}).decompose()
        except:
            pass
        for repeat in article.findAll('section', {'amp-access': 'NOT subscribed AND decision'}):
            repeat.decompose()
        for unwanted in article.findAll(['figure', 'aside', 'amp-ad']):
            unwanted.decompose()
        header_tag = article.find('h1')
        new_header_tag = soup_livemint.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        livemint_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        livemint_html += self.add_print_css()
        livemint_html += '</head><body>' + str(article) + '</body></html>'
        return livemint_html, title

    def insights_url_maker(self, delta=1):
        url_day = datetime.today() - timedelta(days=delta)
        mid_section = url_day.strftime('%Y/%m/%d')
        end_section = url_day.strftime('-%B-%Y')
        url = f'https://www.insightsonindia.com/{mid_section}/insights-daily-current-affairs-pib-summary-{url_day.day}{end_section}/'
        return url


    def insights_new_url_maker(self,delta=1):
        day = datetime.now() - timedelta(days=delta)
        in_day = day.strftime('%d %B %Y')
        req_text = f'INSIGHTS DAILY CURRENT AFFAIRS + PIB SUMMARY- {in_day}'
        base_url = 'https://www.insightsonindia.com/insights-ias-upsc-current-affairs/'
        req_url = self.get_random_response(base_url)
        soup = BeautifulSoup(req_url.content, 'lxml')
        for child in soup.find_all('div', {'class': 'list_div'})[0].ul:
            # following is to account for 2 July and 02 July in text
            if SequenceMatcher(None, child.text,req_text).ratio() > 0.95:
                return child.a['href']

    @http_error
    def parse_insights_daily(self, dummy_url):
        dummy_url = ''
        day = datetime.now()
        if day.strftime('%A') == 'Monday':
            parse_url = self.insights_url_maker(delta=2)
        else:
            parse_url = self.insights_url_maker()
        # print(parse_url)
        html, title = self.parse_outline_url(parse_url)
        return html, title

    @http_error
    def parse_insights_daily_non_outline(self, dummy_url, url_index):
        dummy_url = ''
        day = datetime.now()
        if day.strftime('%A') == 'Monday':
            parse_url = self.insights_new_url_maker(delta=2)
        else:
            parse_url = self.insights_new_url_maker()
        if not parse_url:
            # this is because july table is not updated in early month... did this on 02/7/2020
            parse_url = self.insights_url_maker()
        req_insights = self.get_random_response(parse_url)
        soup = BeautifulSoup(req_insights.content, 'lxml')
        header = soup.find('h1', {'class': 'entry-title'}).text
        header = f'{url_index}_{header}'
        html_insights = f'<html><head><meta charset="utf-8"><title>{soup.title.text}</title>'
        html_insights += self.add_print_css()
        html_insights += f'</head><body><h1><a href="{parse_url}">{header}</a></h1>'
        article = soup.find('div', {'class': 'pf-content'})
        article.findAll('blockquote')[0].decompose()
        for hj in article.findAll('noscript'):
            hj.decompose()
        for figure in article.findAll('img', {'class': 'alignnone'}):
            img_url = figure['data-lazy-src']
            img_response = requests.get(img_url)
            encoded_string = base64.b64encode(img_response.content).decode('ascii')
            figure['class'] = 'center'
            figure['src'] = f"data:image/jpeg;base64,{encoded_string}"
            for key in figure.attrs.copy():
                if key == 'class' or key == 'src':
                    continue
                else:
                    figure.attrs.pop(key)
        html_insights += str(article) + '</body></html>'
        return html_insights, header

    def rebuild_epw(self, call_func=0):
        if call_func == 0:
            print('rebuilding epw article list')
        day = datetime.now()
        temp_week = day.isocalendar()[1]
        epw_week_url = f'https://www.epw.in/journal/{day.isocalendar()[0]}/{temp_week - 1}'
        print(epw_week_url)
        user_agent_epw = 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'
        headers_epw = {'User-Agent': user_agent_epw, 'Referer': 'https://www.facebook.com/',
                       'X-Forwarded-For': '66.249.66.1'}
        epw_response = requests.get(epw_week_url, headers=headers_epw)
        epw_soup = BeautifulSoup(epw_response.content, 'lxml')
        epw_article_list_wrapper = epw_soup.find('div', {'id': 'block-system-main'})
        for trunc1 in epw_article_list_wrapper(['h3']):
            trunc1.decompose()
        epw_article_list = []
        skip_list = ['/engage/', '/appointmentsprogrammesannouncements/', '/current-statistics/']
        for trunc in epw_article_list_wrapper(['a']):
            if '/author/' in trunc['href']:
                trunc.decompose()
            elif any(item in trunc['href'] for item in skip_list):
                continue
            else:
                if call_func == 0:
                    print(f"https://www.epw.in{trunc['href']}")
                epw_article_list.append(f"https://www.epw.in{trunc['href']}")
        i = 0
        if call_func == 1:
            return epw_article_list
        with open('epw.txt', 'w+') as epw_write:
            for url in epw_article_list:
                if i != 0:
                    epw_write.write('\n' + url)
                else:
                    epw_write.write(url)
                i += 1

    def rebuild_economist(self, week_url = None):
        day = datetime.now()
        print('rebuilding economist article list')
        if not week_url:
            print_edition_date = day + timedelta(days=1)
            economist_week_url = 'https://www.economist.com/weeklyedition/' + print_edition_date.strftime('%Y-%m-%d')
        else:
            economist_week_url = week_url
        print(economist_week_url)
        e_response = self.get_random_response(economist_week_url)
        article_list = []
        soup_economist = BeautifulSoup(e_response.content, 'lxml')
        current_edition = soup_economist.find('div', {'class': 'layout-weekly-edition'})
        for item in current_edition(['a'])[2:]:
            if 'graphic-detail/' in item['href'] \
                    or 'economic-and-financial-indicators/' in item['href']:
                continue
            article_list.append(f"https://www.economist.com{item['href']}")
        article_list = list(set(article_list))
        i = 0
        with open('economist.txt', 'w+') as f_write:
            for url in article_list:
                if i != 0:
                    f_write.write('\n' + url)
                else:
                    f_write.write(url)
                i += 1

    def random_economist(self, index=8):
        e_lines = []
        r_lines = []
        with open('economist.txt') as f:
            for line in f:
                e_lines.append(line.strip('\n'))
        if len(e_lines) > 11:
            for itera in range(0, index):
                item = random.choice(e_lines)
                r_lines.append(item)
                e_lines.remove(item)
                itera += 1
        else:
            for item in e_lines:
                r_lines.append(item)
        return r_lines

    def random_epw(self, index=3):
        e_lines = []
        r_lines = []
        with open('epw.txt') as f:
            for line in f:
                e_lines.append(line.strip('\n'))
        if len(e_lines) > 3:
            for iter in range(0, index):
                item = random.choice(e_lines)
                r_lines.append(item)
                e_lines.remove(item)
                iter += 1
        else:
            for item in e_lines:
                r_lines.append(item)
        return r_lines

    def update_economist_epw(self, url_list):
        e_lines = []
        with open('economist.txt') as f:
            for line in f:
                e_lines.append(line.strip('\n'))
        epw_lines = []
        with open('epw.txt') as f:
            for line in f:
                epw_lines.append(line.strip('\n'))
        for url in url_list:
            for url2 in e_lines:
                if SequenceMatcher(None, url, url2).ratio() > 0.85:
                    e_lines.remove(url2)
            for url3 in epw_lines:
                if SequenceMatcher(None, url, url3).ratio() > 0.85:
                    epw_lines.remove(url3)
        print(f'economist articles remaining are {len(e_lines)}')
        print(f'epw article remaining are :{len(epw_lines)}')
        print('updating economist and epw file')
        i = 0
        with open('economist.txt', 'w+') as f_write:
            for url in e_lines:
                if i != 0:
                    f_write.write('\n' + url)
                else:
                    f_write.write(url)
                i += 1
        i = 0
        with open('epw.txt', 'w+') as epw_write:
            for url in epw_lines:
                if i != 0:
                    epw_write.write('\n' + url)
                else:
                    epw_write.write(url)
                i += 1

    def excel_return_urls(self, sh=0):
        url_list = []
        wb = openpyxl.load_workbook('N_Today.xlsx')
        if sh == 0:
            sheet1 = wb['manual']
            url_list = [sheet1.cell(row=x, column=1).value for x in range(1, sheet1.max_row + 1) if
                        sheet1.cell(row=x, column=1).value]
            # for i in range(1, sheet1.max_row + 1):
            #     url_list.append(sheet1.cell(row=i, column=1).value)
        else:
            sheet2 = wb['auto']
            url_list = [sheet2.cell(row=x, column=1).value for x in range(1, sheet2.max_row + 1) if
                        sheet2.cell(row=x, column=1).value]
            # for i in range(1, sheet2.max_row + 1):
            #     url_list.append(sheet2.cell(row=i, column=1).value)
        return url_list

    def update_excel(self, url_list, final=False):
        wb = openpyxl.load_workbook('N_Today.xlsx')
        if not final:
            sheet1 = wb['auto']
            i = 1
            for url in url_list:
                sheet1.cell(row=i, column=1).value = url
                i += 1
            try:
                wb.save('N_Today.xlsx')
            except PermissionError:
                os.system("taskkill /im excel.exe")
                sleep(1)
                wb.save('N_Today.xlsx')
                os.system('start N_Today.xlsx')
        elif final:
            sheet_done = wb['done-urls']
            for url in url_list:
                sheet_done.append([url])
            try:
                wb.save('N_Today.xlsx')
            except PermissionError:
                os.system("taskkill /im excel.exe")
                sleep(1)
                wb.save('N_Today.xlsx')
                os.system('start N_Today.xlsx')
        return True

    @http_error
    def parse_economist(self, input_url, url_index):
        amp_url, title = self.get_amp_url_requests(input_url)
        if not amp_url and not title:
            amp_url, title = self.get_amp_url_selenium(input_url)
        economist_response = self.get_random_response(amp_url)
        soup_economist = BeautifulSoup(economist_response.content, 'lxml')
        eco_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        eco_html += self.add_print_css() + '</head><body>'
        header_tag = soup_economist.find('span', {'class': 'article__headline'})
        new_header_tag = soup_economist.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        eco_html += str(soup_economist.find('header', {'class': 'article__header'}))
        soup_economist.find('div', {'class': "layout-article-links"}).decompose()
        # soup_economist.find('div', {'amp-access':'NOT (NOT cm AND NOT c) AND (
        # loggedIn AND cm.viewsLeft <= 0 AND s.isSubscriber = "false")'}).decompose()
        article = soup_economist.find('div', {'class': 'layout-article-body'})
        for item in article.findAll('figure'):
            item.decompose()
        for item in article.findAll('div', {'class': 'advert'}):
            item.decompose()
        for item in article.findAll('iframe'):
            item.decompose()
        # eco_html += str(soup_economist.find('p', {'class':'article__footnote'}))
        eco_html += str(article)
        eco_html += '</body></html>'
        return eco_html, title

    @http_error
    def parse_hkfp(self, input_url, url_index):
        response = self.get_random_response(input_url)
        soup = BeautifulSoup(response.content, 'lxml')
        header = soup.find('h1', {'class': 'entry-title'})
        del header['class']
        title = re.sub('^\s+', '', soup.title.text)
        html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        html += self.add_print_css() + '</head><body>'
        html += f'<h1><a href="{input_url}">{url_index}_{header.text}</a></h1>'
        article = soup.find('div', {'class': 'entry-content'})
        for item in article.find_all(['figure', 'aside', 'section']):
            item.decompose()
        for item in article.find_all('p'):
            html += str(item)
        html += '</body></html>'
        return html, title

    # noinspection PyTypeChecker
    @http_error
    def parse_indian_express_url(self, input_url):
        # print('indian_express function called')
        regex_search = 'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        response1 = self.get_random_response(input_url)
        soup_func = BeautifulSoup(response1.content, "lxml")
        for sc1 in soup_func(['script', 'style']):
            sc1.decompose()
        for div1 in soup_func.findAll("div", {'class': 'pdsc-related-modify'}):
            div1.decompose()
        for span1 in soup_func.findAll("span", {'class': 'embed-youtube'}):
            span1.decompose()
        article_tag = soup_func.find("div", {"itemprop": "articleBody"})
        html1 = f'<html><head><meta charset="utf-8"><title>{soup_func.title.text}</title>'
        html1 += self.add_print_css() + '</head><body>'
        html1 += f"<h1><a href='{input_url}'>{soup_func.title.text}</a></h1>"
        check_id = 0
        skip_list = ['<img ', '<strong>Opinion', '>Express Explained</', '<strong>Don’t']
        for tag in article_tag:
            count = 0
            if tag.name == "blockquote" and "twitter" in str(tag):
                continue
            if tag.name == 'div':
                try:
                    check = tag['id']
                except:
                    check = tag['class']
                if 'div-gpt-ad' in check or 'id_newsletter_subscription' in check:
                    break
            if tag.name == 'div' and tag['class'][0] == 'share-social':
                # print(tag)
                check_id = 1
            if tag.name == 'span' or tag.name == 'p':
                try:
                    if tag["itemprop"] == "image":
                        continue
                except:
                    if tag.name == "p":
                        if any(x in str(tag) for x in
                               skip_list):  # "<img "not in str(tag) and '<strong>Opinion' not in str(tag) and '<strong>📣' not in str(tag):
                            # print('skipping explained twitter telegram tag')
                            continue
                        else:
                            html1 = html1 + str(tag)
                            continue
                img_urls = re.findall(pattern=regex_search, string=str(tag))
                img_urls = list(set(img_urls))
                if len(img_urls) != 0:
                    for item in img_urls:
                        if item[-4:] == '.jpg' or item[-5:] == '.jpeg':
                            response2 = requests.get(item)
                            # print(item)
                            # print(response2.content)
                            encoded_string = base64.b64encode(response2.content).decode('ascii')
                            html1 += f'<p><img src="data:image/jpeg;base64,{encoded_string}"/></p>'
                            if tag.noscript:
                                html1 += f"<center>{tag.text}</center>"
                            break
                    continue
            if check_id != 1:
                html1 = html1 + str(tag)
            check_id = 0
        html1 += '</body></html>'
        return html1, soup_func.title.text

    @http_error
    def parse_indian_express_faster(self, input_url, url_index):
        response_ie = self.get_random_response(input_url)
        soup_ie = BeautifulSoup(response_ie.content, 'lxml')
        title = soup_ie.title.text
        html_ie = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        html_ie += self.add_print_css() + '</head><body>'
        header = soup_ie.find('h1', {'class': 'native_story_title'})
        html_ie += f"<h1><a href='{input_url}'>{url_index}_{header.text}</a></h1>"
        article = soup_ie.find('div', {'class': 'full-details'})
        article.find('div', {'class': 'share-social'}).decompose()
        for figure in article.findAll('img', {'class': ['size-full', 'size-medium']}):
            figure.decompose()
        for x in article.findAll(['noscript', 'script']):
            x.decompose()
        for z in article.findAll(True, {'class': ['appstext', 'storytags', 'more-from', 'abbott-disc', 'embed-youtube',
                                                  'custom-caption', 'inhouseimg', 'ie-int-campign-ad',
                                                  'pdsc-related-modify']}):
            z.decompose()
        for i in article.findAll(True, {'id': ['id_newsletter_subscription',
                                               'story_content_parts', re.compile('^div-gpt-ad.*')]}):
            i.decompose()
        html_ie += str(article)
        return html_ie, title

    @http_error
    def parse_perspective_anthro(self, input_url, url_index):
        amp_url, title = self.get_amp_url_requests(input_url)
        if not amp_url and not title:
            amp_url, title = self.get_amp_url_selenium(input_url)
        if not amp_url:
            html, title = self.parse_guradian_nytimes_globaltimes_url(input_url)
            return html, title
        response_perspec = self.get_random_response(amp_url)
        soup_perspec = BeautifulSoup(response_perspec.content, 'lxml')
        perspec_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        perspec_html += self.add_print_css() + '</head><body>'
        article = soup_perspec.find('article', {'class': 'amp-wp-article'})
        article.find('footer').decompose()
        article.find('div', {'class': 'sharedaddy'}).decompose()
        article.find('nav', {'data-layout': 'grid'}).decompose()
        for item in article.findAll('amp-img'):
            item.decompose()
        header_tag = article.find('h1', {'class': 'amp-wp-title'})
        new_header_tag = soup_perspec.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        perspec_html += str(article)
        return perspec_html, title

    @http_error
    def parse_sapiens(self, input_url, url_index):
        options = Options()
        fp = webdriver.FirefoxProfile(
            'C:\\Users\\Sabyasachi\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\703g68w9.python_user')
        fp.DEFAULT_PREFERENCES['frozen']['extensions.autoDisableScopes'] = 0
        options.set_preference('extensions.enabledScopes', 15)
        options.add_argument('--headless')
        fp.update_preferences()
        driver = webdriver.Firefox(firefox_profile=fp, options=options)
        driver.get(input_url)
        try:
            element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CLASS_NAME, "entry-content"))
            )
        except:
            pass
        title = driver.title
        soup_sapiens = BeautifulSoup(driver.page_source, 'lxml')
        driver.close()
        sapiens_html = f"<html><head><meta charset='utf-8'><title>{title}</title>"
        sapiens_html += self.add_print_css() + '</head><body>'
        del soup_sapiens.find('h1', {'itemprop': 'headline'})['class']
        header_tag = soup_sapiens.find('h1', {'itemprop': 'headline'})
        new_header_tag = soup_sapiens.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        sapiens_html += f"{str(header_tag)}"
        article = soup_sapiens.find('div', {'class': 'entry-content'})
        for item in soup_sapiens.findAll(['aside', 'figure']):
            item.decompose()
        for ri in soup_sapiens.findAll('div', {'class': 'widget'}):
            ri.decompose()
        sapiens_html += str(article)
        return sapiens_html, title

    @http_error
    def parse_taipei_times(self, input_url, url_index):
        # print(input_url)
        user_agent_wsj = 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'
        headers_wsj = {'User-Agent': user_agent_wsj, 'Referer': 'https://www.facebook.com/',
                       'X-Forwarded-For': '66.249.66.1'}
        response_taiwan = requests.get(input_url, headers=headers_wsj)
        soup_taiwan = BeautifulSoup(response_taiwan.content, 'lxml')
        html = f'<html><head><meta charset="utf-8"><title>{soup_taiwan.title.text}</title>'
        html += self.add_print_css() + '</head><body>'
        article_taiwan = soup_taiwan.find('div', {'class': 'archives'})
        for item in article_taiwan.find_all('div', {'class': ['imgboxa', 'boxTitle']}):
            item.decompose()
        header_tag = soup_taiwan.find('h1')
        new_header_tag = soup_taiwan.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        author_tag = soup_taiwan.find('div', {'class': 'name'})
        # print(author_tag.text)
        new_author_tag = soup_taiwan.new_tag('ul')
        new_author_tag_li = soup_taiwan.new_tag('li')
        new_author_tag_li.string = author_tag.text
        new_author_tag.append(new_author_tag_li)
        for item in article_taiwan(['ul']):
            item.decompose()
        header_tag.insert_after(new_author_tag)
        html += str(article_taiwan) + '</body></html>'
        return html, soup_taiwan.title.text

    @http_error
    def parse_guradian_nytimes_globaltimes_url(self, input_url, url_index):
        url = 'https://mercury.postlight.com/amp?url=' + input_url
        mob_agent = 'Mozilla/5.0 (Linux; Android 10; SM-M315F) AppleWebKit/537.36 ' \
                    '(KHTML, like Gecko) Chrome/81.0.4044.117 Mobile Safari/537.36'
        mob_headers = {'User-Agent': mob_agent}
        response_guardian = requests.get(url, headers=mob_headers)
        if response_guardian.status_code == 404:
            # print('error 404 going the selenium way')
            options = Options()
            fp = webdriver.FirefoxProfile(
                'C:\\Users\\Sabyasachi\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\703g68w9.python_user')
            fp.DEFAULT_PREFERENCES['frozen']['extensions.autoDisableScopes'] = 0
            options.set_preference('extensions.enabledScopes', 15)
            # options.set_preference("browser.link.open_newwindow", 3)
            # options.set_preference("browser.link.open_newwindow.restriction", 0)
            fp.update_preferences()
            driver = webdriver.Firefox(firefox_profile=fp, options=options)
            driver.get(url)
            soup_guardian = BeautifulSoup(driver.page_source, 'lxml')
            driver.close()
        else:
            soup_guardian = BeautifulSoup(response_guardian.content, 'lxml')
        title = soup_guardian.title.text
        head_styles = soup_guardian.findAll('style')
        guardian_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        guardian_html += self.add_print_css()
        for item in head_styles:
            guardian_html += str(item)
        guardian_html += "</head><body>"
        article_full = soup_guardian.find({"article": "hg-article-container"})
        for item in article_full(['figure', 'aside']):
            item.decompose()
        for ki in article_full.find_all('amp-img'):
            ki.decompose()
        article_full.find('div', {'class': 'hg-social-logo-block'}).decompose()
        header_tag = article_full.find('h1', {'class': 'hg-title'})
        new_header_tag = soup_guardian.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        guardian_html += str(article_full)
        guardian_html += "</body></html>"
        return guardian_html, title

    # noinspection PyTypeChecker
    @http_error
    def parse_hindu_url(self, input_hindu_url):
        if 'cartoonscape' in input_hindu_url:
            return None, None
        regex_search = 'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        response_hindu = self.get_random_response(input_hindu_url)
        html_hindu_internal = Pq(response_hindu.content)
        html_content_string = html_hindu_internal('div.article>div:nth-child(3)').html()
        lead_img_html = html_hindu_internal('div.article>div:nth-child(2)>div:nth-child(3)').html()
        title = re.sub('^\s+|\b\s+\Z', '', html_hindu_internal("title")[0].text)
        if 'Mathrubootham' in title:
            return None, None
        html2 = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        html2 += self.add_print_css() + '</head><body>'
        html2 += f"<h1><a href='{input_hindu_url}'>{title}</a></h1>"
        try:
            author_soup = BeautifulSoup(html_hindu_internal('span.author-img-name').html(), 'lxml')
            for item in author_soup.findAll('a', {'class': 'auth-img'}):
                item.decompose()
            html2 += str(author_soup.a)
        except TypeError:
            html2 += f'<h2>Editorial</h2>'
        try:
            url_list = re.findall(pattern=regex_search, string=str(lead_img_html))
            response2 = requests.get(url_list[1])
            encoded_string = base64.b64encode(response2.content).decode('ascii')
            html2 += f'<p><img src="data:image/jpeg;base64,{encoded_string} align= "middle""/></p>'
        except:
            pass
        soup_content = BeautifulSoup(html_content_string, 'lxml')
        for sch in soup_content(['script', 'style']):
            sch.decompose()
        article_tag = soup_content.find('div', id=re.compile("^content-body-\d+-\d+"))

        for tag in article_tag:
            if tag.name == 'div':
                if 'img' in tag['class'][0]:
                    # print(tag['class'])
                    url_list1 = re.findall(pattern=regex_search, string=str(tag))
                    response3 = requests.get(url_list1[0])
                    encoded_string1 = base64.b64encode(response3.content).decode('ascii')
                    html2 += f'<p><img src="data:image/jpeg;base64,{encoded_string1} align= "middle""/></p>'
                if 'also' in tag['class'][0]:
                    continue
            # if 'Also read | ' in str(tag):
            #     continue
            else:
                html2 += str(tag)
        html2 += '</body></html>'
        return html2, title

    @http_error
    def parse_hindu_faster(self, input_url, url_index):
        if 'cartoonscape' in input_url:
            return None, None
        if '/thread/' in input_url:
            print('redirecting thread hindu')
            html_hindu, title = self.parse_other(input_url)
            return html_hindu, title
        response_hindu = self.get_random_response(input_url)
        soup = BeautifulSoup(response_hindu.content, 'lxml')
        title = re.sub('^\s+|', '', soup.title.text)
        if 'Mathrubootham' in title:
            return None, None
        html_hindu = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        html_hindu += self.add_print_css() + '</head><body>'
        # print(soup)
        article = soup.find('div', {'class': 'article'})
        header_tag = article.find('h1', {'class': ['title', 'special-heading', 'headline']})
        new_header_tag = soup.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        article.find_all('div', {'class', ''})[1].decompose()
        article.find_all('div', {'class', ''})[2].decompose()
        for item in article.find_all('img'):
            item.decompose()
        # return str(article), soup.title.text
        class_decompose = ['support-jlm', 'articlebelowtextad', 'media-body', 'subarticlepay', 'dfp-ad',
                           'img-full-width', 'clear']
        for y in article.find_all(True, {'class': class_decompose}):
            y.decompose()
        for u in article.find_all('script'):
            u.decompose()
        html_hindu += str(article) + '</body></html>'
        return html_hindu, title

    @http_error
    def parse_wsj_url(self, input_url, url_index):
        with open('wsj_style.css', 'r',encoding='utf8') as fg:
            style_css = fg.read()
        wsj_url = input_url[0:20] + 'amp/' + input_url[20:]
        user_agent_wsj = 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'
        headers_wsj = {'User-Agent': user_agent_wsj, 'Referer': 'https://www.facebook.com/',
                       'X-Forwarded-For': '66.249.66.1'}
        response_wsj = requests.get(url=wsj_url, headers=headers_wsj)
        soup_wsj = BeautifulSoup(response_wsj.content, 'lxml')
        soup_wsj.find('div', {'class': 'share-bar'}).decompose()
        for item in soup_wsj.find_all('div', {'class': 'media-object'}):
            item.decompose()
        wsj_html = f'<html><head><meta charset="utf-8"><title>{soup_wsj.title.text}</title>'
        wsj_html += f"{style_css}</head><body>"
        wsj_article_full = soup_wsj.find('main', {"id": "main"})
        new_soup_wsj = BeautifulSoup(str(wsj_article_full), 'lxml')
        for ad_wsj in new_soup_wsj.find_all('div', {'class': 'wsj-ad'}):
            ad_wsj.decompose()
        for wsj_stsy in new_soup_wsj(['script', 'style']):
            wsj_stsy.decompose()
        for double_stop in new_soup_wsj.find_all('div', {'amp-access': 'NOT access'}):
            double_stop.decompose()
        header_tag = new_soup_wsj.find('h1', {'class': 'wsj-article-headline'})
        new_header_tag = new_soup_wsj.new_tag("a", href=f'{input_url}')
        new_header_tag.string = f'{url_index}_{header_tag.text}'
        header_tag.string = ''
        header_tag.append(new_header_tag)
        wsj_html += str(new_soup_wsj) + '</body></html>'
        title = re.sub('^\s+', '', soup_wsj.title.text)
        return wsj_html, title

    @http_error
    def parse_wp_url_selenium(self, input_url, url_index):
        """
        this function takes a url input and returns html and title
        :param url_index: html file index to locate in Indian Express folder
        :param input_url: washington post url
        :return: a list of html elements
        """
        options = Options()
        fp = webdriver.FirefoxProfile(
            'C:\\Users\\Sabyasachi\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\703g68w9.python_user')
        fp.DEFAULT_PREFERENCES['frozen']['extensions.autoDisableScopes'] = 0
        options.set_preference('extensions.enabledScopes', 15)
        options.add_argument('-headless')
        # options.set_preference("browser.link.open_newwindow", 3)
        # options.set_preference("browser.link.open_newwindow.restriction", 0)
        fp.update_preferences()
        driver = webdriver.Firefox(firefox_profile=fp, options=options)
        driver.get(input_url)
        try:
            element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CLASS_NAME, "remainder-content"))
            )
            title = driver.title
            elem = driver.find_element_by_class_name("article-body")
            soup = BeautifulSoup(elem.get_attribute('innerHTML'), 'lxml')
        except:
            np = webdriver.FirefoxProfile(
                'C:\\Users\\Sabyasachi\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\wxoz187x.default-release')
            new_driver = webdriver.Firefox(firefox_profile=np)
            new_driver.get(input_url)
            title = new_driver.title
            elem = new_driver.find_element_by_class_name("article-body")
            soup = BeautifulSoup(elem.get_attribute('innerHTML'), 'lxml')
            new_driver.close()
        title = f'{url_index}_{title}'
        washingtonpost_html = f"<html><head><meta charset='utf-8'><title>{title}</title>"
        washingtonpost_html += self.add_print_css() + '</head><body>'
        washingtonpost_html += f"<h1><a href='{input_url}'>{title}</a></h1>"
        for item in soup.find_all('p', {'class': 'font--body'}):
            del item['class']
            washingtonpost_html += str(item)
        washingtonpost_html += '</body></html>'
        driver.close()
        return washingtonpost_html, title

    @http_error
    def parse_wp_url_ampway(self, input_url,url_index):
        amp_url, title = self.get_amp_url_requests(input_url)
        if not amp_url:
            amp_url, title = self.get_amp_url_selenium(input_url)
        if not amp_url:
            html_err_free, title = self.parse_outline_url(input_url=input_url,url_index=url_index)
            return html_err_free, title
        title1 = f'{url_index}_{title}'
        user_agent_google = 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'
        headers_google = {'User-Agent': user_agent_google, 'Referer': 'https://www.facebook.com/',
                          'X-Forwarded-For': '66.249.66.1'}
        response_wp = requests.get(amp_url, headers=headers_google)
        soup_wp = BeautifulSoup(response_wp.content, 'lxml')
        article = soup_wp.find('div', {'class': 'article-body'})
        washingtonpost_html = f"<html><head><meta charset='utf-8'><title>{title}</title>"
        washingtonpost_html += self.add_print_css() + '</head><body>'
        washingtonpost_html += f"<h1><a href='{input_url}'>{title1}</a></h1>"
        try:
            for item in article.find_all('p', {'class': 'font--body'}):
                del item['class']
                washingtonpost_html += str(item)
        except AttributeError:
            n_article = soup_wp.find('div', {'class': 'main'})
            for unwanted in n_article.find_all(True,
                                               {'class': ['ent-ad-mob', 'ent-ad-leaderboard', 'interstitial-link ',
                                                          'ent-raw-container', 'ent-video', 'ent-video-fullwidth']}):
                unwanted.decompose()
            washingtonpost_html += str(n_article)
        washingtonpost_html += '</body></html>'
        return washingtonpost_html, title

    def make_section_pdf(self, tag, pdf_array, index):
        i, j = 0, 0
        pdf_writer = PdfFileWriter()
        for item_pdf in (pdf_array):
            pdf_reader = PdfFileReader(item_pdf.Pdf_Address)
            count = 0
            for page in range(pdf_reader.getNumPages()):
                # Add each page to the writer object
                pdf_writer.addPage(pdf_reader.getPage(page))
                if i == 0:
                    parent = pdf_writer.addBookmark(title=tag, pagenum=0)
                    i = 1
                if count == 0:
                    pdf_writer.addBookmark(title=f'{i}_{item_pdf.List_Index}.{item_pdf.Article_Title}', pagenum=j,
                                           parent=parent)
                    count = 1
                j += 1
            i += 1
            # os.system(f'del {item_pdf[1]}')
        internal_pdf_path = "D:\\UPSC\\UPSC 2020\\Newspaper and others\\Hindu Mint and IE"
        internal_file_name = str(date.today()) + f'_IE_{index}.pdf'
        output_path = os.path.join(internal_pdf_path, internal_file_name)
        with open(output_path, 'wb+') as out:
            pdf_writer.write(out)
        return output_path

    def make_final_pdf(self):
        random.shuffle(self.opinion_list) # to randomize wsj articles, earlier it came in one bunch
        final_list_pdf_pc = [self.make_section_pdf(tag='Chapter 1: Opinion_Articles',
                                                   pdf_array=self.opinion_list, index=0),
                             self.make_section_pdf(tag='Chapter 2: Explained_Articles',
                                                   pdf_array=self.explained_list, index=1),
                             self.make_section_pdf(tag='Chapter 3: Other_Articles',
                                                   pdf_array=self.other_list, index=2),
                             self.make_section_pdf(tag='Chapter 4: Economist and EPW',
                                                   pdf_array=self.economist_list, index=3)]
        pdfmerger = PdfFileMerger()
        for file in final_list_pdf_pc:
            pdfmerger.append(file, import_bookmarks=True)
        with open(os.path.join(self.final_pdf_path, self.final_file_name_pdf), 'wb') as obj:
            pdfmerger.write(obj)
        pdfmerger.close()
        for file in final_list_pdf_pc:
            os.system(f'del "{file}"')
        return True

    def decorate_book_cover(self):
        bf_path = os.getcwd()
        os.chdir('C:\\Users\\Sabyasachi\\Google Drive\\Python Projects\\Daily_Compiler')
        try:
            os.remove('edited.jpg')
        except Exception:
            pass
        image = Image.open('book_cover_image.jpg')
        draw = ImageDraw.Draw(image)
        font = ImageFont.truetype('Handlee-Regular.ttf', size=50)
        (x, y) = (10, 936)
        message = f"DATE:-{str(date.today())}"
        color = 'rgb(0, 0, 0)'  # black color
        draw.text((x, y), message, fill=color, font=font)
        image.save('edited.jpg')
        self.indian_express_epub.set_cover(file_name='edited.jpg', content=open('edited.jpg', 'rb').read())
        os.chdir(bf_path)
        return True

    def add_chapter_array(self, art_list, chap_list, heading):
        i = 0
        for article in art_list:
            chapter = epub.EpubHtml(title=article.Article_Title, file_name=f'{i}_{heading}.xhtml', lang='en')
            with open(article.HTML_Address, 'r', encoding='utf8') as art_html:
                chapter.set_content(art_html.read())
            self.indian_express_epub.add_item(chapter)
            chap_list.append(chapter)
            i += 1
        return True

    def add_chapter_to_nav(self, chapter_array):
        for item in chapter_array:
            self.indian_express_epub.spine.append(item)
        return True

    def make_final_epub(self):
        self.indian_express_epub.set_title(str(date.today()) + "_Indian_Express")
        self.indian_express_epub.set_language('en')
        self.indian_express_epub.add_author('Sabyasachi Sharma')
        self.decorate_book_cover()
        self.add_chapter_array(art_list=self.opinion_list, chap_list=self.opinion_chapters, heading='opinion')
        self.add_chapter_array(art_list=self.explained_list, chap_list=self.explained_chapters, heading='explained')
        self.add_chapter_array(art_list=self.other_list, chap_list=self.other_chapters, heading='other')
        self.add_chapter_array(art_list=self.economist_list, chap_list=self.economist_chapters, heading='economist_epw')
        self.indian_express_epub.toc = (epub.Section('Opinion'), tuple(self.opinion_chapters)), \
                                       (epub.Section('Explained'), tuple(self.explained_chapters)), \
                                       (epub.Section('Other'), tuple(self.other_chapters)), \
                                       (epub.Section('Economist and EPW'), tuple(self.economist_chapters))
        self.indian_express_epub.spine = ['nav']
        self.add_chapter_to_nav(chapter_array=self.opinion_chapters)
        self.add_chapter_to_nav(chapter_array=self.explained_chapters)
        self.add_chapter_to_nav(chapter_array=self.other_chapters)
        self.add_chapter_to_nav(chapter_array=self.economist_chapters)
        self.indian_express_epub.add_item(epub.EpubNcx())
        self.indian_express_epub.add_item(epub.EpubNav())
        os.chdir(self.final_pdf_path)
        epub.write_epub(name=self.epub_file_name, book=self.indian_express_epub)
        return True

    def parse_other(self, input_url, url_index):
        response_other = self.get_random_response(input_url)
        soup_other = BeautifulSoup(response_other.content, 'lxml')
        doc = Article(html=str(soup_other))
        tmp1 = doc.readable
        title = f'{url_index}_{soup_other.title.text}'
        tmp_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
        tmp_html += self.add_print_css() + '</head><body>'
        tmp_html += f"<h1><a href='{input_url}'>{title}</a></h1>" + tmp1 + "</body></html>"
        return tmp_html, title

    def select_parser(self, input_url_host_only, url_full, url_index):
        try:
            func = self.decision_dict[input_url_host_only]
            html, title = func(url_full, url_index)
        except Exception:
            html, title = self.parse_other(url_full, url_index)
        return html, title

    # def make_html_pdf(self,html_string, html_index):
    #     bef_path = os.getcwd()
    #     os.chdir('C:\\Users\\Sabyasachi\\Indian_express_temp')
    #     with open(f'{html_index}.html', 'w+', encoding='utf8') as html_write:
    #         html_write.write(html_string)
    #     self.driver.refresh()
    #     self.driver.get(f'{html_index}.html')
    #
    #     def execute(script, args):
    #         self.driver.execute('executePhantomScript', {'script': script, 'args': args})
    #
    #     self.driver.command_executor._commands['executePhantomScript'] = ('POST', '/session/$sessionId'
    #                                                                          '/phantom/execute')
    #     with open('papersize.js', 'r') as gu:
    #         temp_script = gu.read()
    #     page_format = f'this.{temp_script};'
    #     execute(page_format, [])
    #     render = '''this.render("{}")'''.format(f'{html_index}.pdf')
    #     execute(render, [])
    #     print('pdf generated')
    #     os.chdir(bef_path)
    #     return True

    def parse_history_articles(self, input_url, url_index):
        url_index = 'is_dummy'
        if input_url == 'http://www.thepeoplehistory.com/this-day-in-history.html':
            response_thisday = self.get_random_response(input_url)
            soup_hist = BeautifulSoup(response_thisday.content, 'lxml')
            title = soup_hist.title.text
            for sc_sty_hist in soup_hist(['script', 'style']):
                sc_sty_hist.decompose()
            tmp_html = f'<html><head><meta charset="utf-8"><title>{soup_hist.title.text}</title>' \
                       f'<link rel="stylesheet" href=' \
                       '"http://www.thepeoplehistory.com/style-7.css" type="text/css"><link rel="stylesheet" href=' \
                       '"media-queries.css" type="text/css">'
            tmp_html += self.add_print_css() + '</head><body>'
            hist_tag = soup_hist.find('div', {"id": "left-content"})
            for x in hist_tag(['img', 'small']):
                x.decompose()
            for y in hist_tag(['h2', 'a']):
                if 'This Week In History' in str(y):
                    y.decompose()
            tmp_html += re.sub(pattern='Taken From Our This Day In History From <br/> to <br/>', repl='',
                               string=str(hist_tag))
            tmp_html += '</body></html>'

        elif input_url == 'https://www.indianage.com/indian_history':
            # print('parsing indian age')
            # browser_indianage = webdriver.Firefox(
            #     executable_path='C:\\Users\\Sabyasachi\\Google Drive\\Python Projects\\Daily_Compiler\\geckodriver.exe')
            # browser_indianage.get(input_url)
            # try:
            #     element = WebDriverWait(browser_indianage, 10).until(
            #         EC.presence_of_element_located((By.CLASS_NAME, "container"))
            #     )
            # except:
            #     pass
            url = 'https://www.indianage.com/indian_history'
            user_agent_wsj = 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'
            headers_wsj = {'User-Agent': user_agent_wsj, 'Referer': 'https://www.facebook.com/',
                           'X-Forwarded-For': '66.249.66.1'}
            response_hist = requests.get(input_url, headers=headers_wsj)
            soup = BeautifulSoup(response_hist.content, 'lxml')
            doc = Article(html=str(soup))
            tmp1 = doc.readable
            day = datetime.today().strftime('%B %d')
            title = f'Today in Indian History - Events for {day}'
            tmp_html = f'<html><head><meta charset="utf-8"><title>{title}</title>'
            tmp_html += self.add_print_css() + '</head><body>'
            tmp_html += f"<body><h1><a href='{input_url}'>{title}</a></h1>" + tmp1 + "</body></html>"
            browser_indianage.close()
        return tmp_html, title

    def free_express_folder(self):
        os.chdir('C:\\Users\\Sabyasachi\\Indian_express_temp')
        os.system('del *.') #delete all html with no extensions
        for item in os.listdir('C:\\Users\\Sabyasachi\\Indian_express_temp'):
            if item.endswith(".html") or item.endswith(".pdf"):
                os.remove(item)
        os.chdir('C:\\Users\\Sabyasachi\\Google Drive\\Python Projects\\Daily_Compiler')
        return True
